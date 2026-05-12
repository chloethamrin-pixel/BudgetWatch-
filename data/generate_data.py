"""
BudgetWatch — Data Generation Script
Generates realistic budget vs. actuals data for Contoso Manufacturing.
Produces two Excel files:
  - contoso_budget_raw.xlsx   : messy, unformatted (simulates the "before")
  - contoso_budget_clean.xlsx : clean, structured (what Power BI reads)
"""

import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

DEPARTMENTS = ["Operations", "Marketing", "Human Resources", "IT", "Finance"]

CATEGORIES = [
    "Salaries & Benefits",
    "Travel & Entertainment",
    "Office Supplies",
    "Software & Licenses",
    "Training & Development",
    "Contractor Services",
    "Equipment",
]

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

YEAR = 2024

# Base monthly budgets per department per category (in USD)
BASE_BUDGETS = {
    "Operations": {
        "Salaries & Benefits":    85000,
        "Travel & Entertainment": 12000,
        "Office Supplies":         3000,
        "Software & Licenses":     5000,
        "Training & Development":  4000,
        "Contractor Services":    20000,
        "Equipment":              15000,
    },
    "Marketing": {
        "Salaries & Benefits":    60000,
        "Travel & Entertainment": 18000,
        "Office Supplies":         2000,
        "Software & Licenses":     8000,
        "Training & Development":  3000,
        "Contractor Services":    25000,
        "Equipment":               5000,
    },
    "Human Resources": {
        "Salaries & Benefits":    55000,
        "Travel & Entertainment":  5000,
        "Office Supplies":         1500,
        "Software & Licenses":     4000,
        "Training & Development":  8000,
        "Contractor Services":     6000,
        "Equipment":               2000,
    },
    "IT": {
        "Salaries & Benefits":    75000,
        "Travel & Entertainment":  4000,
        "Office Supplies":         1000,
        "Software & Licenses":    20000,
        "Training & Development":  6000,
        "Contractor Services":    15000,
        "Equipment":              18000,
    },
    "Finance": {
        "Salaries & Benefits":    65000,
        "Travel & Entertainment":  6000,
        "Office Supplies":         2000,
        "Software & Licenses":     7000,
        "Training & Development":  4000,
        "Contractor Services":     8000,
        "Equipment":               3000,
    },
}

def get_variance_multiplier(dept, category, month_index):
    """
    Returns a multiplier applied to budget to get actuals.
    1.0 = exactly on budget. >1.0 = over budget. <1.0 = under budget.
    Encodes realistic patterns:
      - Operations/Marketing overspend in Q3-Q4
      - IT overspends Software & Licenses in Q1 (annual renewals)
      - HR underspends Training in Q1-Q2
      - A few dramatic spikes for alert interest
    """
    base = 1.0 + random.uniform(-0.05, 0.07)  # slight natural noise

    # Q3-Q4 overspend: Operations Travel & Entertainment
    if dept == "Operations" and category == "Travel & Entertainment":
        if month_index >= 6:  # July-December
            base += random.uniform(0.15, 0.35)

    # Q3-Q4 overspend: Marketing Contractor Services
    if dept == "Marketing" and category == "Contractor Services":
        if month_index >= 6:
            base += random.uniform(0.18, 0.30)

    # Q1 IT annual software renewals spike
    if dept == "IT" and category == "Software & Licenses":
        if month_index == 0:  # January
            base += random.uniform(0.40, 0.60)
        elif month_index == 1:  # February (tail of renewals)
            base += random.uniform(0.10, 0.20)

    # HR underspends Training in Q1-Q2
    if dept == "Human Resources" and category == "Training & Development":
        if month_index < 6:
            base -= random.uniform(0.20, 0.40)
            base = max(base, 0.3)

    # Operations Equipment spike in Q4 (year-end purchases)
    if dept == "Operations" and category == "Equipment":
        if month_index >= 9:  # October-December
            base += random.uniform(0.20, 0.45)

    # Finance Contractor Services spike mid-year (audit season)
    if dept == "Finance" and category == "Contractor Services":
        if month_index in [3, 4, 5]:  # April-June
            base += random.uniform(0.15, 0.28)

    # Marketing Travel dramatic spike in September (conference season)
    if dept == "Marketing" and category == "Travel & Entertainment":
        if month_index == 8:  # September
            base += random.uniform(0.30, 0.50)

    return round(base, 4)


def build_records():
    records = []
    for dept in DEPARTMENTS:
        for category in CATEGORIES:
            for month_index, month in enumerate(MONTHS):
                budget = BASE_BUDGETS[dept][category]
                multiplier = get_variance_multiplier(dept, category, month_index)
                actual = round(budget * multiplier, 2)
                variance_amount = round(actual - budget, 2)
                variance_pct = round((actual - budget) / budget, 4)
                records.append({
                    "Month": month,
                    "Month_Number": month_index + 1,
                    "Year": YEAR,
                    "Department": dept,
                    "Category": category,
                    "Budget_Amount": budget,
                    "Actual_Amount": actual,
                    "Variance_Amount": variance_amount,
                    "Variance_Pct": variance_pct,
                })
    return records


def save_clean(df, path):
    df_clean = df.drop(columns=["Month_Number"])
    df_clean.to_excel(path, index=False, sheet_name="Budget_Data")

    wb = load_workbook(path)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [12, 6, 18, 28, 16, 16, 18, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Conditional color on Variance_Pct (column H = index 8)
    red_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF99")
    green_fill = PatternFill(fill_type="solid", fgColor="CCFFCC")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        pct_cell = row[7]  # Variance_Pct column
        try:
            val = float(pct_cell.value)
            if val < -0.10:
                pct_cell.fill = red_fill
            elif val < -0.05:
                pct_cell.fill = yellow_fill
            else:
                pct_cell.fill = green_fill
        except (TypeError, ValueError):
            pass

    wb.save(path)
    print(f"Clean file saved: {path}")


def save_raw(df, path):
    # Messy version: mixed column order, no formatting, includes Month_Number, inconsistent casing
    df_raw = df[["Month_Number", "Month", "Year", "Department", "Category",
                 "Budget_Amount", "Actual_Amount", "Variance_Amount", "Variance_Pct"]].copy()
    df_raw.columns = ["month_num", "Month", "year", "dept", "cost_category",
                      "budget", "actual", "variance", "var_pct"]
    # Add some messiness: duplicate header-like rows every 50 rows
    rows = [df_raw.columns.tolist()]
    for i, row in df_raw.iterrows():
        if i > 0 and i % 50 == 0:
            rows.append(df_raw.columns.tolist())  # repeated header
        rows.append(row.tolist())

    df_messy = pd.DataFrame(rows[1:], columns=rows[0])
    df_messy.to_excel(path, index=False, sheet_name="Sheet1")
    print(f"Raw file saved: {path}")


if __name__ == "__main__":
    records = build_records()
    df = pd.DataFrame(records)

    # Sort by department, category, month for clean ordering
    df = df.sort_values(["Department", "Category", "Month_Number"]).reset_index(drop=True)

    base = "C:/Users/chloe/OneDrive/Desktop/Portfolio/data"
    save_clean(df, f"{base}/contoso_budget_clean.xlsx")
    save_raw(df, f"{base}/contoso_budget_raw.xlsx")

    # Quick summary
    over_budget = df[df["Variance_Pct"] < -0.10]
    print(f"\nTotal rows: {len(df)}")
    print(f"Rows over 10% budget: {len(over_budget)}")
    print("\nWorst variances:")
    print(df.nsmallest(5, "Variance_Pct")[["Department", "Category", "Month", "Variance_Pct"]].to_string(index=False))
